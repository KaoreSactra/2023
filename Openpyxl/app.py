import openpyxl

#criar uma planilha (book)
book = openpyxl.Workbook()
#como visualizar páginas existemtes
print(book.sheetnames)
#Criar uma página
book.create_sheet('Frutas')
#como selecionar uma página
frutas_pages = book['Frutas']
frutas_pages.append(['hdsdoiysidoisad', 'Quantidade', 'Preço unitário'])
frutas_pages.append(['Maçã', '6', 'R$1,60'])
frutas_pages.append(['Laranja', '7', 'R$1,10'])
frutas_pages.append(['Tomate', '3', 'R$2,00'])
frutas_pages.append(['Melancia', '2', 'R$10,00'])

#Salvar planilha
book.save('Planilha de Compras.xlsx')